# analysis_v2.py — 按用户改版样本(122列)重建大乐透分析+样式。先生成样本供核对,确认后再并入 lottery.py。
import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from lottery import get_element_type, get_prop_type, count_consecutive_numbers

BASE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'sample_output')

# 颜色
C_ZONE1 = 'FFFFE0'  # 1-9 浅黄
C_ZONE2 = 'ADD8E6'  # 10-19 浅蓝
C_ZONE3 = '90EE90'  # 20-29 浅绿
C_ZONE4 = 'E6E6FA'  # 30-35 浅紫
C_RED_LOW = 'FFE4B5'   # 红1-红9 浅橙
C_RED_HIGH = 'D9D2E9'  # 红10-红12 浅紫


def diagonal_count(cur, prev):
    """斜号:本期每个号码 n,统计上期中等于 n-1 或 n+1 的个数之和。首期=0。"""
    if prev is None:
        return 0
    ps = set(prev)
    total = 0
    for n in cur:
        if (n - 1) in ps:
            total += 1
        if (n + 1) in ps:
            total += 1
    return total


def repeat_count(cur, prev):
    if prev is None:
        return 0
    ps = set(prev)
    return sum(1 for n in cur if n in ps)


def same_position_count(nums):
    """同位:末位(%10)重复的个数(超过首次出现的部分)。"""
    return int(pd.Series([n % 10 for n in nums]).duplicated().sum())


def cross_props(prev, cur):
    """跨期相克:对每个球位,用上期同位球的五行 + 本期五行 查表。首期全 None。"""
    if prev is None:
        return [None] * len(cur)
    return [get_prop_type(get_element_type(prev[i]) + get_element_type(cur[i])) for i in range(len(cur))]


def dlt_header():
    h = [(' ', '期数'), (' ', '号码'), (' ', '和值')]
    for i in range(1, 6):
        h.append(('蓝%d' % i, '五行%d' % i))
        h.append(('蓝%d' % i, '相克%d' % i))
    # 余数 one-hot(保留原始标签笔误:蓝4 为 4-0,4-2,4-2)
    rem_labels = {
        1: ['1-0', '1-1', '1-2'], 2: ['2-0', '2-1', '2-2'], 3: ['3-0', '3-1', '3-2'],
        4: ['4-0', '4-2', '4-2'], 5: ['5-0', '5-1', '5-2'],
    }
    for i in range(1, 6):
        for lab in rem_labels[i]:
            h.append(('蓝%d' % i, lab))
    for lab in ['1-9', '10-19', '20-29', '30-35', '重号', '斜号', '连号', '同位']:
        h.append(('蓝球统计', lab))
    h.append(('奇偶', '奇'))
    h.append(('奇偶', '偶'))
    for i in range(1, 6):
        h.append(('蓝%d' % i, '奇'))
        h.append(('蓝%d' % i, '偶'))
    for i in range(1, 6):
        h.append(('蓝%d' % i, '大'))
        h.append(('蓝%d' % i, '小'))
    for n in range(1, 10):
        h.append(('蓝球1区', str(n)))
    for n in range(10, 20):
        h.append(('蓝球2区', str(n)))
    for n in range(20, 36):
        h.append(('蓝球3区', str(n)))
    h.append(('红球', '后区号码'))
    h.append(('红球', '红和值'))
    h.append(('红1', '红五行1'))
    h.append(('红1', '红相克1'))
    h.append(('红2', '红五行2'))
    h.append(('红2', '红相克2'))
    for lab in ['1_0', '1_1', '1_2']:
        h.append(('红1', lab))
    for lab in ['2_0', '2_1', '2_2']:
        h.append(('红2', lab))
    h.append(('红奇偶', '红奇数'))
    h.append(('红奇偶', '红偶数'))
    for i in range(1, 13):
        h.append(('红球分布', '红%d' % i))
    h.append(('红球分布', '一区'))
    h.append(('红球分布', '二区'))
    return h


def build_dlt_row(period, haoma, fronts, backs, prev_f, prev_b):
    row = []
    row.append(period)
    row.append(haoma)
    row.append(sum(fronts))
    f_elem = [get_element_type(n) for n in fronts]
    f_prop = cross_props(prev_f, fronts)
    for i in range(5):
        row.append(f_elem[i])
        row.append(f_prop[i])
    # 前区余数 one-hot(按位置:第 i 球 3 列,余数 r 放第 r 列)
    for i in range(5):
        r = fronts[i] % 3
        cells = [None, None, None]
        cells[r] = r
        row.extend(cells)
    # 蓝球统计
    row.append(sum(1 for n in fronts if n < 10))
    row.append(sum(1 for n in fronts if 9 < n < 20))
    row.append(sum(1 for n in fronts if 19 < n < 30))
    row.append(sum(1 for n in fronts if 29 < n < 40))
    row.append(repeat_count(fronts, prev_f))
    row.append(diagonal_count(fronts, prev_f))
    row.append(count_consecutive_numbers(fronts))
    row.append(same_position_count(fronts))
    # 整体奇偶(计数)
    row.append(sum(1 for n in fronts if n % 2 == 1))
    row.append(sum(1 for n in fronts if n % 2 == 0))
    # 每球位 奇/偶 1-0 标记
    for i in range(5):
        odd = 1 if fronts[i] % 2 == 1 else 0
        row.append(odd)
        row.append(1 - odd)
    # 每球位 大/小 1-0 标记(个位 5-9 大,0-4 小)
    for i in range(5):
        big = 1 if (fronts[i] % 10) >= 5 else 0
        row.append(big)
        row.append(1 - big)
    # 前区分布 one-hot 1..35
    fs = set(fronts)
    for n in range(1, 36):
        row.append(n if n in fs else None)
    # 后区
    row.append('+'.join('%02d' % b for b in backs))  # 后区号码 A+B
    row.append(sum(backs))                            # 红和值
    b_elem = [get_element_type(n) for n in backs]
    b_prop = cross_props(prev_b, backs)
    row.append(b_elem[0]); row.append(b_prop[0])
    row.append(b_elem[1]); row.append(b_prop[1])
    # 后区余数 one-hot
    for i in range(2):
        r = backs[i] % 3
        cells = [None, None, None]
        cells[r] = r
        row.extend(cells)
    # 后区奇偶计数
    row.append(sum(1 for n in backs if n % 2 == 1))
    row.append(sum(1 for n in backs if n % 2 == 0))
    # 红球分布 红1..红12
    bs = set(backs)
    for n in range(1, 13):
        row.append(n if n in bs else None)
    # 一区(0-9 个数)/二区(其余)
    row.append(sum(1 for n in backs if n <= 9))
    row.append(sum(1 for n in backs if n >= 10))
    return row


def daletou_v2(df):
    header = dlt_header()
    # 解析号码
    rows = []
    prev_f = prev_b = None
    fronts_list, backs_list = [], []
    for _, rec in df.iterrows():
        haoma = rec['号码']
        front_str, back_str = haoma.split('+')
        fronts = [int(x) for x in front_str.split('.')]
        backs = [int(x) for x in back_str.split('.')]
        fronts_list.append(fronts); backs_list.append(backs)
    for idx in range(len(df)):
        period = df.iloc[idx]['期数']
        haoma = df.iloc[idx]['号码']
        fronts = fronts_list[idx]
        backs = backs_list[idx]
        pf = fronts_list[idx - 1] if idx > 0 else None
        pb = backs_list[idx - 1] if idx > 0 else None
        r = build_dlt_row(period, haoma, fronts, backs, pf, pb)
        assert len(r) == len(header), f"row {len(r)} != header {len(header)}"
        rows.append(r)

    df_final = pd.DataFrame(rows)
    df_final.columns = pd.MultiIndex.from_tuples(header)

    os.makedirs(BASE_DIR, exist_ok=True)
    temp = os.path.join(BASE_DIR, '_tmp_v2.xlsx')
    df_final.to_excel(temp)
    wb = load_workbook(temp)
    ws = wb.active

    n_cols = df_final.shape[1] + 1  # +1 for index col A
    # 相克字体色
    _dye_value(['↑克', '↓克'], 'FF0000', ws, n_cols)
    _dye_value(['刑'], '00FF00', ws, n_cols)
    _dye_value(['↑生', '↓生'], '0000FF', ws, n_cols)
    # 前区分区底色(按 level1 数字标签)
    _dye_zone([str(n) for n in range(1, 10)], C_ZONE1, ws)
    _dye_zone([str(n) for n in range(10, 20)], C_ZONE2, ws)
    _dye_zone([str(n) for n in range(20, 30)], C_ZONE3, ws)
    _dye_zone([str(n) for n in range(30, 36)], C_ZONE4, ws)
    # 红球分布着色:红1-红9 / 红10-红12
    _dye_group_labels('红球分布', [f'红{i}' for i in range(1, 10)], C_RED_LOW, ws)
    _dye_group_labels('红球分布', [f'红{i}' for i in range(10, 13)], C_RED_HIGH, ws)

    # 边框 + 居中 + 表头加粗
    thin = Side(style='thin', color='000000')
    base_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = base_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
    for cell in ws[2]:
        cell.font = Font(bold=True)

    # 加粗分组竖线(按第一层表头连续分组的左右边界)
    _bold_group_dividers(ws, header)

    # 列宽(按表头标签动态匹配)
    _set_widths(ws, df_final.shape[1], header)
    return wb


def _dye_value(value_list, color, ws, max_col):
    font = Font(color=color)
    for row in ws.iter_rows(min_row=3, min_col=1, max_col=max_col):
        for cell in row:
            if cell.value in value_list:
                cell.font = font


def _dye_zone(dye_range, color, ws):
    cols = []
    for cell in ws[2]:
        if cell.value in dye_range:
            cols.append(cell.column)
    fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    for col in cols:
        for row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
            for cell in row:
                cell.fill = fill


def _dye_group_labels(group_name, labels, color, ws):
    """在指定一级分组(row1)下,给 row2 标签在 labels 里的列上色。"""
    # 找出属于该 group 的列(用合并单元格或重复值判断)。简单起见:row2 标签匹配且 row1 链接到该 group。
    fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    # 建立列->group 映射
    group_at = {}
    for mr in ws.merged_cells.ranges:
        if mr.min_row == 1:
            g = ws.cell(row=1, column=mr.min_col).value
            for c in range(mr.min_col, mr.max_col + 1):
                group_at[c] = g
    for c in range(1, ws.max_column + 1):
        g = group_at.get(c, ws.cell(row=1, column=c).value)
        lab = ws.cell(row=2, column=c).value
        if g == group_name and lab in labels:
            for row in ws.iter_rows(min_row=2, min_col=c, max_col=c):
                for cell in row:
                    cell.fill = fill


def _bold_group_dividers(ws, header):
    thick = Side(style='thick', color='000000')
    thin = Side(style='thin', color='000000')
    # header 第 0 元素对应 Excel 第 2 列(第 1 列是 index)
    groups = [g for g, _ in header]
    # 连续分组的列区间(Excel 列号,从 2 起)
    runs = []
    start = 0
    for i in range(1, len(groups) + 1):
        if i == len(groups) or groups[i] != groups[start]:
            runs.append((start + 2, i - 1 + 2))  # excel col range
            start = i
    for (c1, c2) in runs:
        for r in range(1, ws.max_row + 1):
            left = ws.cell(row=r, column=c1)
            b = left.border
            left.border = Border(left=thick, right=b.right, top=b.top, bottom=b.bottom)
            right = ws.cell(row=r, column=c2)
            b2 = right.border
            right.border = Border(left=b2.left, right=thick, top=b2.top, bottom=b2.bottom)


def _set_widths(ws, ncols, header):
    wide6 = set()
    wide7 = set()
    w_houqu = set()
    for i, (grp, lab) in enumerate(header):
        col = i + 2
        if grp.startswith('蓝') and (lab.startswith('五行') or lab.startswith('相克')):
            wide6.add(col)                       # 前区 五行/相克
        elif grp == '蓝球统计':
            wide6.add(col)                       # 蓝球统计全部 8 列
        elif grp in ('红1', '红2') and (lab.startswith('红五行') or lab.startswith('红相克')):
            wide6.add(col)                       # 后区 五行/相克
        elif grp == '红球' and lab == '后区号码':
            w_houqu.add(col)                     # 后区号码 A+B
        elif grp == '红球' and lab == '红和值':
            wide7.add(col)
    for col in range(1, ncols + 2):
        letter = ws.cell(column=col, row=2).column_letter
        if col == 2:                      # 期数
            ws.column_dimensions[letter].width = 8
        elif col == 3:                    # 号码
            ws.column_dimensions[letter].width = 23
        elif col == 4:                    # 和值
            ws.column_dimensions[letter].width = 6
        elif col in w_houqu:              # 后区号码 A+B
            ws.column_dimensions[letter].width = 10
        elif col in wide7:                # 红和值
            ws.column_dimensions[letter].width = 7
        elif col in wide6:                # 五行/相克、蓝球统计
            ws.column_dimensions[letter].width = 6
        else:
            ws.column_dimensions[letter].width = 3.5


def ssq_header():
    """双色球:6 红球(标 蓝1~蓝6) + 1 蓝球(标 红球/后区)。套用大乐透同款新增列。"""
    h = [(' ', '期数'), (' ', '号码'), (' ', '和值')]
    for i in range(1, 7):
        h.append(('蓝%d' % i, '五行%d' % i))
        h.append(('蓝%d' % i, '相克%d' % i))
    rem_labels = {
        1: ['1-0', '1-1', '1-2'], 2: ['2-0', '2-1', '2-2'], 3: ['3-0', '3-1', '3-2'],
        4: ['4-0', '4-2', '4-2'], 5: ['5-0', '5-1', '5-2'], 6: ['6-0', '6-1', '6-2'],
    }
    for i in range(1, 7):
        for lab in rem_labels[i]:
            h.append(('蓝%d' % i, lab))
    for lab in ['1-9', '10-19', '20-29', '30-33', '重号', '斜号', '连号', '同位']:
        h.append(('蓝球统计', lab))
    h.append(('奇偶', '奇'))
    h.append(('奇偶', '偶'))
    for i in range(1, 7):
        h.append(('蓝%d' % i, '奇'))
        h.append(('蓝%d' % i, '偶'))
    for i in range(1, 7):
        h.append(('蓝%d' % i, '大'))
        h.append(('蓝%d' % i, '小'))
    for n in range(1, 10):
        h.append(('蓝球1区', str(n)))
    for n in range(10, 20):
        h.append(('蓝球2区', str(n)))
    for n in range(20, 34):
        h.append(('蓝球3区', str(n)))
    # 后区(单蓝球)
    h.append(('红球', '后区号码'))
    h.append(('红球', '红五行'))
    h.append(('红球', '红相克'))
    for lab in ['1_0', '1_1', '1_2']:
        h.append(('红1', lab))
    # 红球分布 红1..红16
    for i in range(1, 17):
        h.append(('红球分布', '红%d' % i))
    h.append(('红球分布', '一区'))
    h.append(('红球分布', '二区'))
    return h


def build_ssq_row(period, haoma, reds, blue, prev_r, prev_b):
    row = [period, haoma, sum(reds)]
    r_elem = [get_element_type(n) for n in reds]
    r_prop = cross_props(prev_r, reds)
    for i in range(6):
        row.append(r_elem[i])
        row.append(r_prop[i])
    for i in range(6):
        r = reds[i] % 3
        cells = [None, None, None]
        cells[r] = r
        row.extend(cells)
    row.append(sum(1 for n in reds if n < 10))
    row.append(sum(1 for n in reds if 9 < n < 20))
    row.append(sum(1 for n in reds if 19 < n < 30))
    row.append(sum(1 for n in reds if 29 < n < 40))
    row.append(repeat_count(reds, prev_r))
    row.append(diagonal_count(reds, prev_r))
    row.append(count_consecutive_numbers(reds))
    row.append(same_position_count(reds))
    row.append(sum(1 for n in reds if n % 2 == 1))
    row.append(sum(1 for n in reds if n % 2 == 0))
    for i in range(6):
        odd = 1 if reds[i] % 2 == 1 else 0
        row.append(odd)
        row.append(1 - odd)
    for i in range(6):
        big = 1 if (reds[i] % 10) >= 5 else 0
        row.append(big)
        row.append(1 - big)
    rs = set(reds)
    for n in range(1, 34):
        row.append(n if n in rs else None)
    # 后区单蓝球
    row.append('%02d' % blue)               # 后区号码(单个,不用 A+B)
    row.append(get_element_type(blue))       # 红五行
    # 红相克:跨期与上一期蓝球比
    row.append(get_prop_type(get_element_type(prev_b) + get_element_type(blue)) if prev_b is not None else None)
    r = blue % 3
    cells = [None, None, None]
    cells[r] = r
    row.extend(cells)
    # 红球分布 红1..红16
    for n in range(1, 17):
        row.append(n if n == blue else None)
    # 一区(1-9)/二区(10-16)
    row.append(1 if blue <= 9 else 0)
    row.append(1 if blue >= 10 else 0)
    return row


def shuangseqiu_v2(df):
    header = ssq_header()
    reds_list, blue_list = [], []
    for _, rec in df.iterrows():
        front_str, back_str = rec['号码'].split('+')
        reds_list.append([int(x) for x in front_str.split('.')])
        blue_list.append(int(back_str))
    rows = []
    for idx in range(len(df)):
        pr = reds_list[idx - 1] if idx > 0 else None
        pb = blue_list[idx - 1] if idx > 0 else None
        r = build_ssq_row(df.iloc[idx]['期数'], df.iloc[idx]['号码'],
                          reds_list[idx], blue_list[idx], pr, pb)
        assert len(r) == len(header), f"ssq row {len(r)} != header {len(header)}"
        rows.append(r)
    df_final = pd.DataFrame(rows)
    df_final.columns = pd.MultiIndex.from_tuples(header)

    os.makedirs(BASE_DIR, exist_ok=True)
    temp = os.path.join(BASE_DIR, '_tmp_ssq_v2.xlsx')
    df_final.to_excel(temp)
    wb = load_workbook(temp)
    ws = wb.active
    n_cols = df_final.shape[1] + 1
    _dye_value(['↑克', '↓克'], 'FF0000', ws, n_cols)
    _dye_value(['刑'], '00FF00', ws, n_cols)
    _dye_value(['↑生', '↓生'], '0000FF', ws, n_cols)
    _dye_zone([str(n) for n in range(1, 10)], C_ZONE1, ws)
    _dye_zone([str(n) for n in range(10, 20)], C_ZONE2, ws)
    _dye_zone([str(n) for n in range(20, 30)], C_ZONE3, ws)
    _dye_zone([str(n) for n in range(30, 34)], C_ZONE4, ws)
    _dye_group_labels('红球分布', [f'红{i}' for i in range(1, 10)], C_RED_LOW, ws)
    _dye_group_labels('红球分布', [f'红{i}' for i in range(10, 17)], C_RED_HIGH, ws)
    thin = Side(style='thin', color='000000')
    base_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = base_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
    for cell in ws[2]:
        cell.font = Font(bold=True)
    _bold_group_dividers(ws, header)
    _set_widths_ssq(ws, df_final.shape[1], header)
    return wb


def _set_widths_ssq(ws, ncols, header):
    # 按表头(group,label)动态确定需加宽的列(列号 = header索引 + 2,因首列是 index)
    wide6 = set()   # 宽度 6 的列
    wide7 = set()   # 宽度 7 的列
    for i, (grp, lab) in enumerate(header):
        col = i + 2
        if grp.startswith('蓝') and (lab.startswith('五行') or lab.startswith('相克')):
            wide6.add(col)                       # 前区 五行/相克
        elif grp == '蓝球统计':
            wide6.add(col)                       # 蓝球统计全部 8 列(含连号/同位)
        elif grp == '红球' and lab in ('后区号码', '红五行', '红相克'):
            wide7.add(col)                       # 后区号码/五行/相克
    for col in range(1, ncols + 2):
        letter = ws.cell(column=col, row=2).column_letter
        if col == 2:
            ws.column_dimensions[letter].width = 8
        elif col == 3:
            ws.column_dimensions[letter].width = 23
        elif col == 4:
            ws.column_dimensions[letter].width = 6
        elif col in wide7:
            ws.column_dimensions[letter].width = 7
        elif col in wide6:
            ws.column_dimensions[letter].width = 6
        else:
            ws.column_dimensions[letter].width = 3.5


if __name__ == '__main__':
    import random
    random.seed(7)
    rows = []
    for i in range(12):
        fronts = sorted(random.sample(range(1, 36), 5))
        backs = sorted(random.sample(range(1, 13), 2))
        h = '.'.join('%02d' % x for x in fronts) + '+' + '.'.join('%02d' % x for x in backs)
        rows.append([f'{24001 + i}', h])
    df = pd.DataFrame(rows, columns=['期数', '号码'])
    wb = daletou_v2(df)
    out = os.path.join(BASE_DIR, '大乐透_v2样本.xlsx')
    wb.save(out)
    print('已生成:', out)

    # 双色球样本
    random.seed(11)
    rows = []
    for i in range(12):
        reds = sorted(random.sample(range(1, 34), 6))
        blue = random.randint(1, 16)
        h = '.'.join('%02d' % x for x in reds) + '+' + '%02d' % blue
        rows.append([f'{24001 + i}', h])
    df = pd.DataFrame(rows, columns=['期数', '号码'])
    wb = shuangseqiu_v2(df)
    out = os.path.join(BASE_DIR, '双色球_v2样本.xlsx')
    wb.save(out)
    print('已生成:', out)
