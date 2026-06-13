import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill,Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl import Workbook
import os
import requests
from lxml import etree

desk = os.path.join(os.path.expanduser("~"), "Desktop")
try:
    os.makedirs(os.path.join(desk, 'lottery'), exist_ok=True)
except OSError:
    pass
BASE_DIR = os.path.join(desk, 'lottery')

def transform1(nums):
    return [int(x) for x in nums.split('+')[0].split('.')]


def transform2(nums):
    return int(nums.split('+')[1])

def transform3(nums):
    return [int(x) for x in nums.split('+')[1].split('.')]

def process(num):
    last_digit = num % 10
    return last_digit


def process2(num):
    last_digit = num % 3
    return last_digit  


def get_element_type(num):
    last_digit = num % 10
    if last_digit in [0, 5]:
        return "土"
    elif last_digit in [1, 6]:
        return "水"
    elif last_digit in [2, 7]:
        return "火"
    elif last_digit in [3, 8]:
        return "木"
    elif last_digit in [4, 9]:
        return "金"


def get_prop_type(item):
    if item in ['金金', '木木', '水水', '火火', '土土']:
        return "刑"
    elif item in ['金木', '木土', '土水', '水火', '火金']:
        return "↓克"
    elif item in ['木金', '土木', '水土', '火水', '金火']:
        return "↑克"
    elif item in ['金水', '水木', '木火', '火土', '土金']:
        return "↓生"
    elif item in ['水金', '木水', '火木', '土火', '金土']:
        return "↑生"


def expand_list(row, max_num):
    s = pd.Series(index=range(max_num + 1))
    s[row] = row
    return s


def expand_list_(row, max_num):
    s = pd.Series(index=range(1, max_num + 1))
    for i in row:
        s[row] = row
    return s


def process_yushu(df):
    final = []
    for i in range(df.shape[1]):
        df_ = df.iloc[:, i].explode().apply(expand_list, max_num=2)
        df_.columns = [str(i + 1) + '-' + str(_) for _ in df_.columns]
        final.append(df_)
    return pd.concat(final, axis=1)

def process_yushu2(df):
    final = []
    for i in range(df.shape[1]):
        df_ = df.iloc[:, i].explode().apply(expand_list, max_num=2)
        df_.columns = [str(i + 1) + '_' + str(_) for _ in df_.columns]
        final.append(df_)
    return pd.concat(final, axis=1)

def count_consecutive_numbers(nums):
    count = 0
    i = 0
    while i < len(nums) - 1:
        if nums[i] + 1 == nums[i + 1]:
            while i < len(nums) - 1 and nums[i] + 1 == nums[i + 1]:
                i += 1
            count += 1
        i += 1
    return count


def count_repeat_numbers(matrix):
    repeats_per_row = [0]

    # 遍历每一行，统计该行及上一行的数字是否有重复
    for i in range(1, matrix.shape[0]):
        repeats = np.sum(np.isin(matrix[i], matrix[i - 1]))
        repeats_per_row.append(repeats)

    return repeats_per_row


def count_same_numbers(x):
    return sum(pd.Series(x).duplicated())


def dye_column(dye_range, color, ws):
    dye_columns = []
    for cell in ws[2]:
        if cell.value in dye_range:
            # 保存默认边框设置
            dye_columns.append(cell.column)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    for col in dye_columns:
        for row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
            for cell in row:
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")


def dye_value(value_list, color, ws ,max_col):
    font = Font(color=color)
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=max_col):
        for cell in row:
            if cell.value in value_list:
                cell.font = font

def modi_1(df1, df2):
    col1 = df1.columns.tolist()
    col2 = df2.columns.tolist()
    
    # 生成交替的列名
    new_col = [item for pair in zip(col1, col2) for item in pair]
    
    # 确保 df_new 只有 new_col 中的列
    df_new = pd.concat([df1, df2], axis=1)
    
    # 只选择存在于 df_new 中的列
    new_col = [col for col in new_col if col in df_new.columns]
    
    return df_new[new_col]


# 创建多级列索引
multi_index = pd.MultiIndex.from_tuples([(' ', '期数'), (' ', '号码'), (' ', '和值'),
                                         (' ', '五行1'), (' ', '相克1'), (' ', '五行2'),
                                         (' ', '相克2'), (' ', '五行3'), (' ', '相克3'),
                                         (' ', '五行4'), (' ', '相克4'), (' ', '五行5'),
                                         (' ', '相克5'), (' ', '五行6'), (' ', '相克6'),
                                         ('蓝1', '1-0'), ('蓝1', '1-1'), ('蓝1', '1-2'),
                                         ('蓝2', '2-0'), ('蓝2', '2-1'), ('蓝2', '2-2'),
                                         ('蓝3', '3-0'), ('蓝3', '3-1'), ('蓝3', '3-2'),
                                         ('蓝4', '4-0'), ('蓝4', '4-2'), ('蓝4', '4-2'),
                                         ('蓝5', '5-0'), ('蓝5', '5-1'), ('蓝5', '5-2'),
                                         ('蓝6', '6-0'), ('蓝6', '6-1'), ('蓝6', '6-2'),
                                         ('奇偶', '奇数'), ('奇偶', '偶数'), ('蓝球1区', '1'),
                                         ('蓝球1区', '2'), ('蓝球1区', '3'), ('蓝球1区', '4'),
                                         ('蓝球1区', '5'), ('蓝球1区', '6'), ('蓝球1区', '7'),
                                         ('蓝球1区', '8'), ('蓝球1区', '9'), ('蓝球2区', '10'),
                                         ('蓝球2区', '11'), ('蓝球2区', '12'), ('蓝球2区', '13'),
                                         ('蓝球2区', '14'), ('蓝球2区', '15'), ('蓝球2区', '16'),
                                         ('蓝球2区', '17'), ('蓝球2区', '18'), ('蓝球2区', '19'),
                                         ('蓝球3区', '20'), ('蓝球3区', '21'), ('蓝球3区', '22'),
                                         ('蓝球3区', '23'), ('蓝球3区', '24'), ('蓝球3区', '25'),
                                         ('蓝球3区', '26'), ('蓝球3区', '27'), ('蓝球3区', '28'),
                                         ('蓝球3区', '29'), ('蓝球3区', '30'), ('蓝球3区', '31'),
                                         ('蓝球3区', '32'), ('蓝球3区', '33'), ('蓝球统计', '1-9'),
                                         ('蓝球统计', '10-19'),('蓝球统计', '20-29'),('蓝球统计', '30-33'),('蓝球统计', '连号'),
                                         ('蓝球统计', '重号'), ('蓝球统计', '同位'), ('红球', '红球'),
                                         ('红球', '红球五行'), ('红球', '红球相克'), ('红球分布', '红1'),
                                         ('红球分布', '红2'), ('红球分布', '红3'), ('红球分布', '红4'),
                                         ('红球分布', '红5'), ('红球分布', '红6'), ('红球分布', '红7'),
                                         ('红球分布', '红8'), ('红球分布', '红9'), ('红球分布', '红10'),
                                         ('红球分布', '红11'), ('红球分布', '红12'), ('红球分布', '红13'),
                                         ('红球分布', '红14'), ('红球分布', '红15'), ('红球分布', '红16'),
                                         ('红球统计', ' 0'), ('红球统计', ' 1'), ('红球统计', ' 2')],
                                        names=[' ', ' '])

# 创建多级列索引
multi_index2 = pd.MultiIndex.from_tuples([(' ', '期数'), (' ', '号码'), (' ', '和值'),
                                         (' ', '五行1'), (' ', '相克1'), (' ', '五行2'),
                                         (' ', '相克2'), (' ', '五行3'),
                                         (' ', '相克3'), (' ', '五行4'), (' ', '相克4'),
                                         (' ', '五行5'), (' ', '相克5'),
                                         ('蓝1', '1-0'), ('蓝1', '1-1'), ('蓝1', '1-2'),
                                         ('蓝2', '2-0'), ('蓝2', '2-1'), ('蓝2', '2-2'),
                                         ('蓝3', '3-0'), ('蓝3', '3-1'), ('蓝3', '3-2'),
                                         ('蓝4', '4-0'), ('蓝4', '4-2'), ('蓝4', '4-2'),
                                         ('蓝5', '5-0'), ('蓝5', '5-1'), ('蓝5', '5-2'),
                                         ('奇偶', '奇数'), ('奇偶', '偶数'), ('蓝球1区', '1'),
                                         ('蓝球1区', '2'), ('蓝球1区', '3'), ('蓝球1区', '4'),
                                         ('蓝球1区', '5'), ('蓝球1区', '6'), ('蓝球1区', '7'),
                                         ('蓝球1区', '8'), ('蓝球1区', '9'), ('蓝球2区', '10'),
                                         ('蓝球2区', '11'), ('蓝球2区', '12'), ('蓝球2区', '13'),
                                         ('蓝球2区', '14'), ('蓝球2区', '15'), ('蓝球2区', '16'),
                                         ('蓝球2区', '17'), ('蓝球2区', '18'), ('蓝球2区', '19'),
                                         ('蓝球3区', '20'), ('蓝球3区', '21'), ('蓝球3区', '22'),
                                         ('蓝球3区', '23'), ('蓝球3区', '24'), ('蓝球3区', '25'),
                                         ('蓝球3区', '26'), ('蓝球3区', '27'), ('蓝球3区', '28'),
                                         ('蓝球3区', '29'), ('蓝球3区', '30'), ('蓝球3区', '31'),
                                         ('蓝球3区', '32'), ('蓝球3区', '33'), ('蓝球3区', '34'), ('蓝球3区', '35'),
                                         ('蓝球统计', '1-9'), ('蓝球统计', '10-19'),('蓝球统计', '20-29'),('蓝球统计', '30-35'),('蓝球统计', '连号'),
                                         ('蓝球统计', '重号'), ('蓝球统计', '同位'), ('红球', '红1'),
                                          ('红球', '红2'),('红球', '红和值'),('红球', '红五行1'),('红球', '红五行2'),
                                          ('红球', '红相克1'), ('红球', '红相克2'),
                                          ('红1', '1_0'), ('红1', '1_1'), ('红1', '1_2'),
                                          ('红2', '2_0'), ('红2', '2_1'), ('红2', '2_2'),('红奇偶', '红奇数'), ('红奇偶', '红偶数'),('红球分布', '红1'),
                                         ('红球分布', '红2'), ('红球分布', '红3'), ('红球分布', '红4'),
                                         ('红球分布', '红5'), ('红球分布', '红6'), ('红球分布', '红7'),
                                         ('红球分布', '红8'), ('红球分布', '红9'), ('红球分布', '红10'),
                                         ('红球分布', '红11'), ('红球分布', '红12'),('红球统计', '红连号'), ('红球统计', '红重号'), ('红球统计', '红同位')],
                                        names=[' ', ' '])


def shuangseqiu(df):
    df['nums1'], df['nums2'] = df.iloc[:, 1].apply(transform1), df.iloc[:, 1].apply(transform2)
    df['和值'] = df['nums1'].apply(sum)
    df['prop'] = np.vectorize(get_element_type)(df['nums1'].tolist()).tolist()
    df['remainder'] = np.vectorize(process)(df['nums1'].tolist()).tolist()
    matrix1 = np.insert(df['prop'].tolist(), 0, '木', axis=0)
    matrix2 = np.insert(df['prop'].tolist(), len(df['nums1'].tolist()), '木', axis=0)
    df['compare'] = np.vectorize(get_prop_type)(np.char.add(matrix1, matrix2)[:-1, :].tolist()).tolist()
    df['余数'] = np.vectorize(process2)(df['nums1'].tolist()).tolist()
    df_1 = df['nums1'].apply(pd.Series)
    df_1.columns = ['数' + str(i) for i in range(1, 7)]
    df_2 = df['prop'].apply(pd.Series)
    df_2.columns = ['五行' + str(i) for i in range(1, 7)]
    df_3 = df['compare'].apply(pd.Series)
    df_3.columns = ['相克' + str(i) for i in range(1, 7)]
    df_3.loc[0, :] = np.nan
    df_4 = df['余数'].apply(pd.Series)
    df_4.columns = ['余数' + str(i) for i in range(1, 7)]
    df_4 = process_yushu(df_4)
    df_4['奇数'] = df['nums1'].apply(lambda x: sum([_ % 2 == 1 for _ in x]))
    df_4['偶数'] = df['nums1'].apply(lambda x: sum([_ % 2 == 0 for _ in x]))
    df_5 = df['nums1'].apply(expand_list_, max_num=33)
    df_5['1-9'] = df['nums1'].apply(lambda x: sum(np.array(x)<10))
    df_5['10-19'] = df['nums1'].apply(lambda x: sum((np.array(x) > 9)&(np.array(x) < 20)))
    df_5['20-29'] = df['nums1'].apply(lambda x: sum((np.array(x) > 19)&(np.array(x) < 30)))
    df_5['30-33'] = df['nums1'].apply(lambda x: sum((np.array(x)> 29)&(np.array(x) < 40)))
    df_5['连号'] = df['nums1'].apply(count_consecutive_numbers)
    df_5['重号'] = count_repeat_numbers(df['nums1'])
    df_5['同位'] = df['remainder'].apply(lambda x: pd.Series(x).duplicated().sum())
    df_5['红球'] = df['nums2'].copy()
    df_5['红球五行'] = np.vectorize(get_element_type)(df['nums2'].tolist()).tolist()
    matrix1 = np.insert(df_5['红球五行'].tolist(), 0, '木', axis=0)
    matrix2 = np.insert(df_5['红球五行'].tolist(), len(df_5['红球五行'].tolist()), '木', axis=0)
    df_5['红球相克'] = np.vectorize(get_prop_type)(np.char.add(matrix1, matrix2)[:-1].tolist()).tolist()
    df_5['红球相克'].iloc[0] = np.nan
    df_6 = df['nums2'].apply(lambda x: expand_list_([x], max_num=16))
    df_6.columns = ['红' + str(x) for x in df_6.columns]
    df['红球余数'] = np.vectorize(process2)(df['nums2'].tolist())
    df_7 = df['红球余数'].apply(expand_list, max_num=2)
    df_7.columns = [' ' + str(x) for x in df_7.columns]
    df_2 = modi_1(df_2, df_3)
    df_final = pd.concat([df.iloc[:, :2], df[['和值']], df_2, df_4, df_5, df_6, df_7], axis=1)
    df_final.columns = [str(x) for x in df_final.columns]
    df_final = df_final.reset_index(drop=True)

    # 修改DataFrame的列索引为多级索引
    df_final.columns = multi_index

    temp_file = 'sample.xlsx'
    df_final.to_excel(os.path.join(BASE_DIR, temp_file))
    # 使用 openpyxl 处理 Excel 文件
    wb = load_workbook(os.path.join(BASE_DIR, temp_file))
    ws = wb.active
    dye_value(['↑克', '↓克'], 'FF0000',ws, df_final.shape[1])
    dye_value(['刑'], '00FF00',ws,df_final.shape[1])
    dye_value(['↑生', '↓生'], '0000FF',ws,df_final.shape[1])

    dye_column([str(_) for _ in np.arange(1, 10)], "FFFFE0",ws)
    dye_column([str(_) for _ in np.arange(10, 20)], "ADD8E6",ws)
    dye_column([str(_) for _ in np.arange(20, 30)], "90EE90",ws)
    dye_column([str(_) for _ in np.arange(30, 34)], "E6E6FA",ws)

    # 创建一个黑色边框
    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000'))

    # 给所有单元格添加黑色边框
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border

    # 设置单元格中的文字居中
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 遍历从A到CP的列并设置宽度为5
    for col in range(1, df_final.shape[1] + 5):  # CP对应的列号为88
        column = ws.cell(column=col, row=2)
        if col == 2:
            ws.column_dimensions[column.column_letter].width = 8
        elif col == 3:
            ws.column_dimensions[column.column_letter].width = 23
        elif col == 4:
            ws.column_dimensions[column.column_letter].width = 6
        elif col in np.arange(5, 17):
            ws.column_dimensions[column.column_letter].width = 6
        elif col in np.arange(70, 80):
            ws.column_dimensions[column.column_letter].width = 6
        else:
            ws.column_dimensions[column.column_letter].width = 3

    return wb

def daletou(df):
    df['nums1'], df['nums2'] = df.iloc[:, 1].apply(transform1), df.iloc[:, 1].apply(transform3)
    df['和值'] = df['nums1'].apply(sum)
    df['prop'] = np.vectorize(get_element_type)(df['nums1'].tolist()).tolist()      ######五行
    df['remainder'] = np.vectorize(process)(df['nums1'].tolist()).tolist() ######尾数
    matrix1 = np.insert(df['prop'].tolist(), 0, '木', axis=0)
    matrix2 = np.insert(df['prop'].tolist(), len(df['nums1'].tolist()), '木', axis=0)
    df['compare'] = np.vectorize(get_prop_type)(np.char.add(matrix1, matrix2)[:-1, :].tolist()).tolist()
    df['余数'] = np.vectorize(process2)(df['nums1'].tolist()).tolist()
    df_1 = df['nums1'].apply(pd.Series)
    df_1.columns = ['数' + str(i) for i in range(1, 6)]
    df_2 = df['prop'].apply(pd.Series)
    df_2.columns = ['五行' + str(i) for i in range(1, 6)]
    df_3 = df['compare'].apply(pd.Series)
    df_3.columns = ['相克' + str(i) for i in range(1, 6)]
    df_3.loc[0, :] = np.nan
    df_4 = df['余数'].apply(pd.Series)
    df_4.columns = ['余数' + str(i) for i in range(1, 6)]
    df_4 = process_yushu(df_4)
    df_4['奇数'] = df['nums1'].apply(lambda x: sum([_ % 2 == 1 for _ in x]))
    df_4['偶数'] = df['nums1'].apply(lambda x: sum([_ % 2 == 0 for _ in x]))
    df_5 = df['nums1'].apply(expand_list_, max_num=35)
    df_5['1-9'] = df['nums1'].apply(lambda x: sum(np.array(x)<10))
    df_5['10-19'] = df['nums1'].apply(lambda x: sum((np.array(x) > 9)&(np.array(x) < 20)))
    df_5['20-29'] = df['nums1'].apply(lambda x: sum((np.array(x) > 19)&(np.array(x) < 30)))
    df_5['30-35'] = df['nums1'].apply(lambda x: sum((np.array(x)> 29)&(np.array(x) < 40)))
    df_5['连号'] = df['nums1'].apply(count_consecutive_numbers)
    df_5['重号'] = count_repeat_numbers(df['nums1'])
    df_5['同位'] = df['remainder'].apply(lambda x: pd.Series(x).duplicated().sum())
    df_2 = modi_1(df_2, df_3)
    df_final = pd.concat([df.iloc[:, :2], df[['和值']], df_2, df_4, df_5], axis=1).copy()

    ############红球######################################
    df['prop'] = np.vectorize(get_element_type)(df['nums2'].tolist()).tolist()      ######五行
    df['remainder'] = np.vectorize(process)(df['nums2'].tolist()).tolist() ######尾数
    matrix1 = np.insert(df['prop'].tolist(), 0, '木', axis=0)
    matrix2 = np.insert(df['prop'].tolist(), len(df['nums2'].tolist()), '木', axis=0)
    df['compare'] = np.vectorize(get_prop_type)(np.char.add(matrix1, matrix2)[:-1, :].tolist()).tolist()
    df['余数'] = np.vectorize(process2)(df['nums2'].tolist()).tolist()
    df_1 = df['nums2'].apply(pd.Series)
    df_1.columns = ['红' + str(i) for i in range(1, 3)]
    df_1['红和值'] = df['nums2'].apply(sum)
    df_2 = df['prop'].apply(pd.Series)
    df_2.columns = ['红五行' + str(i) for i in range(1, 3)]
    df_3 = df['compare'].apply(pd.Series)
    df_3.columns = ['红相克' + str(i) for i in range(1, 3)]
    df_3.loc[0, :] = np.nan
    df_4 = df['余数'].apply(pd.Series)
    df_4.columns = ['红余数' + str(i) for i in range(1, 3)]
    df_4 = process_yushu2(df_4)
    df_4['红奇数'] = df['nums2'].apply(lambda x: sum([_ % 2 == 1 for _ in x]))
    df_4['红偶数'] = df['nums2'].apply(lambda x: sum([_ % 2 == 0 for _ in x]))
    df_5 = df['nums2'].apply(expand_list_, max_num=12)
    df_5.columns = ['红' + str(x) for x in df_5.columns]
    df_5['红连号'] = df['nums2'].apply(count_consecutive_numbers)
    df_5['红重号'] = count_repeat_numbers(df['nums2'])
    df_5['红同位'] = df['remainder'].apply(lambda x: pd.Series(x).duplicated().sum())
    ############################################################
    df_final = pd.concat([df_final, df_1,  df_2, df_3, df_4, df_5], axis=1)
    df_final.columns = [str(x) for x in df_final.columns]
    df_final = df_final.reset_index(drop=True)

    # 修改DataFrame的列索引为多级索引
    df_final.columns = multi_index2
    # print(df_final.columns)
    temp_file = 'sample.xlsx'
    df_final.to_excel(os.path.join(BASE_DIR, temp_file))
    # 使用 openpyxl 处理 Excel 文件
    wb = load_workbook(os.path.join(BASE_DIR, temp_file))
    ws = wb.active
    dye_value(['↑克', '↓克'], 'FF0000',ws, df_final.shape[1])
    dye_value(['刑'], '00FF00',ws,df_final.shape[1])
    dye_value(['↑生', '↓生'], '0000FF',ws,df_final.shape[1])

    dye_column([str(_) for _ in np.arange(1, 10)], "FFFFE0",ws)
    dye_column([str(_) for _ in np.arange(10, 20)], "ADD8E6",ws)
    dye_column([str(_) for _ in np.arange(20, 30)], "90EE90",ws)
    dye_column([str(_) for _ in np.arange(30, 36)], "E6E6FA",ws)

    # 创建一个黑色边框
    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000'))

    # 给所有单元格添加黑色边框
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border

    # 设置单元格中的文字居中
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 遍历从A到CP的列并设置宽度为5
    for col in range(1, df_final.shape[1] + 5):  # CP对应的列号为88
        column = ws.cell(column=col, row=2)
        if col == 2:
            ws.column_dimensions[column.column_letter].width = 8
        elif col == 3:
            ws.column_dimensions[column.column_letter].width = 23
        elif col == 4:
            ws.column_dimensions[column.column_letter].width = 6
        elif col in np.arange(5, 15):
            ws.column_dimensions[column.column_letter].width = 6
        elif col in np.arange(15, 30):
            ws.column_dimensions[column.column_letter].width = 4
        elif col in np.arange(67, df_final.shape[1]+2):
            ws.column_dimensions[column.column_letter].width = 6
        else:
            ws.column_dimensions[column.column_letter].width = 3

    return wb

# if __name__ == "__main__":
# type = input("请选择要处理的类型： 双色球请输入0  大乐透请输入1")

# while True:
#     if type == "0":
#         input_name = '双色球号码.xlsx'
#         output_name = '双色球转化结果.xlsx'
#         df = pd.read_excel(os.path.join(BASE_DIR, input_name), sheet_name='Sheet1', engine='openpyxl')
#         df.columns.values[0] = '期数'
#         df.columns.values[1] = '号码'
#         wb = shuangseqiu(df)
#         wb.save(os.path.join(BASE_DIR, output_name))
#         break
#     elif type == "1":
#         input_name = '大乐透号码.xlsx'
#         output_name = '大乐透转化结果.xlsx'
#         df = pd.read_excel(os.path.join(BASE_DIR, input_name), sheet_name='Sheet1', engine='openpyxl')
#         df.columns.values[0] = '期数'
#         df.columns.values[1] = '号码'
#         wb = daletou(df)
#         wb.save(os.path.join(BASE_DIR, output_name))
#         break
#     else:
#         type = input("选择错误 请选择要处理的类型： 双色球请输入0  大乐透请输入1")

if __name__ == "__main__":
    output_name = '双色球转化结果.xlsx'
    output_name2 = '双色球.csv'
    data = []
    url = "http://datachart.500.com/ssq/history/newinc/history.php?start=00001&end=27001"
    response = requests.get(url)
    response = response.text
    selector = etree.HTML(response)
    for i in selector.xpath('//tr[@class="t_tr1"]'):
        datetime = i.xpath('td/text()')[0]
        red = i.xpath('td/text()')[1:7]
        blue = i.xpath('td/text()')[7:8]
        if datetime != '1':
            data.append([datetime, red, blue])
        
    df = pd.DataFrame(data, columns=['期数', 'nums1', 'nums2'])
    df['号码'] = df.apply(lambda row: '.'.join(map(str, row['nums1'])) + '+' + '.'.join(map(str, row['nums2'])), axis=1)
    df = df[['期数','号码']]
    df = df.iloc[::-1].reset_index(drop=True)
    df.to_csv(os.path.join(BASE_DIR, output_name2), encoding='utf-8-sig', index=False)

    df = df.tail(180).reset_index(drop=True)
    wb = shuangseqiu(df)
    wb.save(os.path.join(BASE_DIR, output_name))



############daletou###########################
    output_name = '大乐透转化结果.xlsx'
    output_name2 = '大乐透.csv'
    data = []
    url = "http://datachart.500.com/dlt/history/newinc/history.php?start=00001&end=27001"
    response = requests.get(url)
    response = response.text
    selector = etree.HTML(response)
    for i in selector.xpath('//tr[@class="t_tr1"]'):
        datetime = i.xpath('td/text()')[0]
        red = i.xpath('td/text()')[1:6]
        blue = i.xpath('td/text()')[6:8]
        if datetime != '1':
            data.append([datetime, red, blue])
        
    df = pd.DataFrame(data, columns=['期数', 'nums1', 'nums2'])
    df['号码'] = df.apply(lambda row: '.'.join(map(str, row['nums1'])) + '+' + '.'.join(map(str, row['nums2'])), axis=1)
    df = df[['期数','号码']]
    df = df.iloc[::-1].reset_index(drop=True)
    df.to_csv(os.path.join(BASE_DIR, output_name2), encoding='utf-8-sig', index=False)

    df = df.tail(180).reset_index(drop=True)
    wb = daletou(df)
    wb.save(os.path.join(BASE_DIR, output_name))