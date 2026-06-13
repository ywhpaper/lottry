# 不联网验证 app_main:用 mock 数据替换 scrape,跑完整流程,检查 5 个产出文件。
import os
import random
import pandas as pd
import app_main


def mock_ssq(*a, **k):
    random.seed(1)
    rows = []
    for i in range(220):
        reds = sorted(random.sample(range(1, 34), 6))
        blue = random.randint(1, 16)
        h = '.'.join('%02d' % x for x in reds) + '+' + '%02d' % blue
        rows.append([str(24001 + i), h])
    return pd.DataFrame(rows, columns=['期数', '号码'])


def mock_dlt(*a, **k):
    random.seed(2)
    rows = []
    for i in range(220):
        fronts = sorted(random.sample(range(1, 36), 5))
        backs = sorted(random.sample(range(1, 13), 2))
        h = '.'.join('%02d' % x for x in fronts) + '+' + '.'.join('%02d' % x for x in backs)
        rows.append([str(24001 + i), h])
    return pd.DataFrame(rows, columns=['期数', '号码'])


# patch scrape to return ssq/dlt mock based on url
_orig = app_main.scrape
def fake_scrape(url, n_red, n_blue):
    return mock_ssq() if n_red == 6 else mock_dlt()
app_main.scrape = fake_scrape

app_main.run()

out = app_main.base_output_dir()
print('\n=== 产出文件检查 ===')
for f in ['双色球.csv', '大乐透.csv', '双色球转化结果.xlsx', '大乐透转化结果.xlsx']:
    p = os.path.join(out, f)
    print(f, '存在' if os.path.exists(p) else '缺失', os.path.getsize(p) if os.path.exists(p) else '')

# verify csv full length and xlsx recent rows + reco
import openpyxl
for name, label in [('双色球转化结果.xlsx', '双色球'), ('大乐透转化结果.xlsx', '大乐透')]:
    wb = openpyxl.load_workbook(os.path.join(out, name))
    ws = wb.active
    print(f'\n{label}: 列数={ws.max_column} 总行数={ws.max_row}')
    # find reco rows
    for r in range(ws.max_row - 6, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(2, 4)]
        if any(vals):
            print('  ', vals)
