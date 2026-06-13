# app_main.py — 彩票分析 exe 主程序。
# 抓取 500.com 历史 -> 全量存 CSV -> 最近180期转化为带样式 xlsx(v2格式)-> 末尾附3注频率加权推荐。
# 产出目录:exe 同目录下的 lottery/ 文件夹。
import os
import sys
import random
import datetime
import requests
from lxml import etree
import pandas as pd
from openpyxl.styles import Font, Alignment

import analysis_v2 as v2


def base_output_dir():
    """exe 同目录(打包后)或脚本同目录(开发时)下的 lottery 文件夹。"""
    if getattr(sys, 'frozen', False):
        here = os.path.dirname(sys.executable)
    else:
        here = os.path.dirname(os.path.abspath(__file__))
    out = os.path.join(here, 'lottery')
    os.makedirs(out, exist_ok=True)
    return out


HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
                  '(KHTML, like Gecko) Chrome/120.0 Safari/537.36'
}
SSQ_URL = 'https://datachart.500.com/ssq/history/newinc/history.php'
DLT_URL = 'https://datachart.500.com/dlt/history/newinc/history.php'


def scrape(url, n_red, n_blue):
    resp = requests.get(url, headers=HEADERS, timeout=30)
    resp.encoding = 'utf-8'
    selector = etree.HTML(resp.text)
    data = []
    for i in selector.xpath('//tr[@class="t_tr1"]'):
        tds = i.xpath('td/text()')
        period = tds[0]
        red = tds[1:1 + n_red]
        blue = tds[1 + n_red:1 + n_red + n_blue]
        if period != '1' and len(red) == n_red and len(blue) == n_blue:
            data.append([period, red, blue])
    df = pd.DataFrame(data, columns=['期数', 'nums1', 'nums2'])
    df['号码'] = df.apply(
        lambda r: '.'.join(map(str, r['nums1'])) + '+' + '.'.join(map(str, r['nums2'])), axis=1)
    df = df[['期数', '号码']]
    df = df.iloc[::-1].reset_index(drop=True)  # 升序
    return df


# ---------- 频率加权推荐 ----------
def _weighted_sample(items, weights, k):
    pool = list(items)
    w = list(weights)
    out = []
    for _ in range(k):
        if not pool:
            break
        total = sum(w)
        x = random.random() * total
        idx = 0
        while idx < len(w) - 1 and x >= w[idx]:
            x -= w[idx]
            idx += 1
        out.append(pool[idx])
        del pool[idx]
        del w[idx]
    return sorted(out)


def _freq_weights(lo, hi, picks):
    counts = {n: 0 for n in range(lo, hi + 1)}
    for arr in picks:
        for n in arr:
            if lo <= n <= hi:
                counts[n] += 1
    items = list(range(lo, hi + 1))
    weights = [counts[n] + 1 for n in items]
    return items, weights


def recommend_ssq(reds_list, blue_list, n=3):
    items, weights = _freq_weights(1, 33, reds_list)
    bi, bw = _freq_weights(1, 16, [[b] for b in blue_list])
    out = []
    for _ in range(n):
        reds = _weighted_sample(items, weights, 6)
        blue = _weighted_sample(bi, bw, 1)[0]
        out.append('红 ' + ' '.join('%02d' % x for x in reds) + '  蓝 %02d' % blue)
    return out


def recommend_dlt(fronts_list, backs_list, n=3):
    fi, fw = _freq_weights(1, 35, fronts_list)
    bi, bw = _freq_weights(1, 12, backs_list)
    out = []
    for _ in range(n):
        fronts = _weighted_sample(fi, fw, 5)
        backs = _weighted_sample(bi, bw, 2)
        out.append('前区 ' + ' '.join('%02d' % x for x in fronts) +
                   '  后区 ' + ' '.join('%02d' % x for x in backs))
    return out


def append_reco(wb, recos, ncols):
    """在工作表数据下方空两行,写推荐号码。"""
    ws = wb.active
    start = ws.max_row + 3
    title = ws.cell(row=start, column=2, value='推荐号码(频率加权,仅供娱乐,不构成投注建议)')
    title.font = Font(bold=True)
    for k, line in enumerate(recos, 1):
        ws.cell(row=start + k, column=2, value='第%d注' % k).font = Font(bold=True)
        ws.cell(row=start + k, column=3, value=line)
    return wb


def parse_ssq(df):
    reds_list, blue_list = [], []
    for h in df['号码']:
        f, b = h.split('+')
        reds_list.append([int(x) for x in f.split('.')])
        blue_list.append(int(b))
    return reds_list, blue_list


def parse_dlt(df):
    fronts_list, backs_list = [], []
    for h in df['号码']:
        f, b = h.split('+')
        fronts_list.append([int(x) for x in f.split('.')])
        backs_list.append([int(x) for x in b.split('.')])
    return fronts_list, backs_list


def run():
    out_dir = base_output_dir()
    v2.BASE_DIR = out_dir  # 让 v2 的临时文件也落在输出目录
    random.seed(datetime.datetime.now().timestamp())

    log = lambda m: print(m, flush=True)

    # ---- 双色球 ----
    log('正在抓取双色球历史数据...')
    ssq = scrape(SSQ_URL, 6, 1)
    ssq.to_csv(os.path.join(out_dir, '双色球.csv'), encoding='utf-8-sig', index=False)
    log('双色球共 %d 期,已存 CSV。生成转化结果(最近180期)...' % len(ssq))
    ssq_recent = ssq.tail(180).reset_index(drop=True)
    wb = v2.shuangseqiu_v2(ssq_recent)
    reds_list, blue_list = parse_ssq(ssq.tail(100))
    recos = recommend_ssq(reds_list, blue_list, 3)
    append_reco(wb, recos, None)
    wb.save(os.path.join(out_dir, '双色球转化结果.xlsx'))
    log('双色球完成。推荐:')
    for r in recos:
        log('  ' + r)

    # ---- 大乐透 ----
    log('正在抓取大乐透历史数据...')
    dlt = scrape(DLT_URL, 5, 2)
    dlt.to_csv(os.path.join(out_dir, '大乐透.csv'), encoding='utf-8-sig', index=False)
    log('大乐透共 %d 期,已存 CSV。生成转化结果(最近180期)...' % len(dlt))
    dlt_recent = dlt.tail(180).reset_index(drop=True)
    wb = v2.daletou_v2(dlt_recent)
    fronts_list, backs_list = parse_dlt(dlt.tail(100))
    recos = recommend_dlt(fronts_list, backs_list, 3)
    append_reco(wb, recos, None)
    wb.save(os.path.join(out_dir, '大乐透转化结果.xlsx'))
    log('大乐透完成。推荐:')
    for r in recos:
        log('  ' + r)

    log('\n全部完成!文件已生成在:%s' % out_dir)


if __name__ == '__main__':
    try:
        run()
    except Exception as e:
        import traceback
        print('运行出错:', e)
        traceback.print_exc()
    input('\n按回车键退出...')
